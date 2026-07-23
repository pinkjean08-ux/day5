from __future__ import annotations

import argparse
import logging
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_URL = "https://www.starbucks.co.kr"
SIDO_URL = f"{BASE_URL}/store/getSidoList.do"
STORE_URL = f"{BASE_URL}/store/getStore.do"
SOURCE_PAGE = f"{BASE_URL}/store/store_map.do"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/131.0.0.0 Safari/537.36"
    ),
    "Referer": SOURCE_PAGE,
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "X-Requested-With": "XMLHttpRequest",
}

OUTPUT_COLUMNS = [
    ("매장코드", "s_code"),
    ("매장명", "s_name"),
    ("영문매장명", "s_ename"),
    ("시도코드", "sido_code"),
    ("시도", "sido_name"),
    ("시군구코드", "gugun_code"),
    ("시군구", "gugun_name"),
    ("도로명주소", "doro_address"),
    ("지번주소", "address"),
    ("전화번호", "tel"),
    ("위도", "lat"),
    ("경도", "lot"),
    ("개점일", "open_dt"),
    ("매장유형", "store_type"),
    ("매장상세URL", "store_url"),
]


def build_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)
    return session


def post_json(
    session: requests.Session,
    url: str,
    data: dict[str, Any],
    retries: int = 4,
    timeout: int = 30,
) -> dict[str, Any]:
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            response = session.post(url, data=data, timeout=timeout)
            response.raise_for_status()
            payload = response.json()
            if not isinstance(payload, dict):
                raise ValueError("JSON 응답이 객체 형식이 아닙니다.")
            return payload
        except (requests.RequestException, ValueError) as exc:
            last_error = exc
            if attempt == retries:
                break
            wait = 2 ** (attempt - 1)
            logging.warning("요청 실패(%s/%s): %s; %s초 후 재시도", attempt, retries, exc, wait)
            time.sleep(wait)
    raise RuntimeError(f"스타벅스 서버 요청에 실패했습니다: {last_error}")


def get_sido_list(session: requests.Session) -> list[dict[str, Any]]:
    payload = post_json(session, SIDO_URL, {})
    sido_list = payload.get("list", [])
    if not sido_list:
        raise RuntimeError("시도 목록이 비어 있습니다. 공식 사이트의 응답 형식이 변경되었을 수 있습니다.")
    return sido_list


def get_stores_by_sido(
    session: requests.Session,
    sido_code: str,
    iend: int = 3000,
) -> list[dict[str, Any]]:
    data = {
        "ins_lat": "37.56682",
        "ins_lng": "126.97865",
        "p_sido_cd": sido_code,
        "p_gugun_cd": "",
        "in_biz_cd": "",
        "set_date": "",
        "iend": str(iend),
        "searchType": "C",
        "all_store": "0",
    }
    payload = post_json(session, STORE_URL, data)
    stores = payload.get("list", [])
    if stores is None:
        return []
    if not isinstance(stores, list):
        raise RuntimeError(f"시도코드 {sido_code}의 매장 응답 형식이 예상과 다릅니다.")
    return stores


def normalize_store(store: dict[str, Any]) -> dict[str, Any]:
    row = dict(store)
    for key in ("lat", "lot"):
        try:
            row[key] = float(row[key]) if row.get(key) not in (None, "") else None
        except (TypeError, ValueError):
            row[key] = None

    open_dt = str(row.get("open_dt") or "").strip()
    if len(open_dt) == 8 and open_dt.isdigit():
        row["open_dt"] = f"{open_dt[:4]}-{open_dt[4:6]}-{open_dt[6:]}"

    biz_code = str(row.get("s_code") or row.get("s_seq") or "").strip()
    row["store_url"] = f"{SOURCE_PAGE}?in_biz_cd={biz_code}" if biz_code else SOURCE_PAGE

    store_type_values = []
    flag_labels = {
        "store_type": "",
        "defaultimage": "",
        "my_siren_order_store_yn": "사이렌오더",
        "drive_through": "드라이브스루",
    }
    existing = str(row.get("store_type") or "").strip()
    if existing:
        store_type_values.append(existing)
    for key, label in flag_labels.items():
        if label and str(row.get(key) or "").upper() in {"Y", "1", "TRUE"}:
            store_type_values.append(label)
    row["store_type"] = ", ".join(dict.fromkeys(store_type_values))
    return row


def collect_all(delay: float = 0.35) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    session = build_session()
    sido_list = get_sido_list(session)
    all_stores: list[dict[str, Any]] = []

    for index, sido in enumerate(sido_list, start=1):
        code = str(sido.get("sido_cd") or "").strip()
        name = str(sido.get("sido_nm") or code).strip()
        if not code:
            continue
        stores = get_stores_by_sido(session, code)
        logging.info("[%s/%s] %s: %s개", index, len(sido_list), name, len(stores))
        all_stores.extend(normalize_store(store) for store in stores)
        time.sleep(delay)

    unique: dict[str, dict[str, Any]] = {}
    for row in all_stores:
        key = str(row.get("s_code") or "").strip()
        if not key:
            key = "|".join(
                str(row.get(k) or "").strip()
                for k in ("s_name", "doro_address", "lat", "lot")
            )
        unique[key] = row

    stores = list(unique.values())
    stores.sort(key=lambda r: (
        str(r.get("sido_name") or ""),
        str(r.get("gugun_name") or ""),
        str(r.get("s_name") or ""),
    ))
    return stores, sido_list


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws, max_width: int = 55) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 10), max_width)


def write_excel(stores: list[dict[str, Any]], output_path: Path, collected_at: datetime) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "전국매장"
    ws.append([title for title, _ in OUTPUT_COLUMNS])
    for store in stores:
        ws.append([store.get(key, "") for _, key in OUTPUT_COLUMNS])
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["H"].width = 48
    ws.column_dimensions["I"].width = 48
    ws.column_dimensions["O"].width = 55
    for col in ("K", "L"):
        for cell in ws[col][1:]:
            cell.number_format = "0.000000"
    for cell in ws["O"][1:]:
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

    summary = wb.create_sheet("시군구별요약")
    summary.append(["시도", "시군구", "매장수"])
    counts = Counter(
        (str(s.get("sido_name") or ""), str(s.get("gugun_name") or ""))
        for s in stores
    )
    for (sido, gugun), count in sorted(counts.items()):
        summary.append([sido, gugun, count])
    style_header(summary)
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = summary.dimensions
    summary.sheet_view.showGridLines = False
    auto_width(summary)

    metadata = wb.create_sheet("수집정보")
    metadata_rows = [
        ["항목", "내용"],
        ["수집일시(KST)", collected_at.strftime("%Y-%m-%d %H:%M:%S")],
        ["수집매장수", len(stores)],
        ["출처", SOURCE_PAGE],
        ["수집방식", "스타벅스 코리아 매장찾기 페이지가 사용하는 JSON 응답을 시도별로 요청"],
        ["주의사항", "공식 공개 API가 아니므로 사이트 구조 변경 시 코드 수정이 필요할 수 있음"],
        ["연구사용", "분석 전 폐점·중복·주소 행정구역 및 수집일 기준을 반드시 점검"],
    ]
    for row in metadata_rows:
        metadata.append(row)
    style_header(metadata)
    metadata.sheet_view.showGridLines = False
    metadata.column_dimensions["A"].width = 22
    metadata.column_dimensions["B"].width = 95
    for cell in metadata["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    metadata["B4"].hyperlink = SOURCE_PAGE
    metadata["B4"].style = "Hyperlink"

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="전국 스타벅스 매장 정보를 엑셀로 저장합니다.")
    parser.add_argument("--output", default="output/starbucks_korea_stores.xlsx")
    parser.add_argument("--delay", type=float, default=0.35, help="시도별 요청 사이 대기시간(초)")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    collected_at = datetime.now().astimezone()
    stores, _ = collect_all(delay=max(args.delay, 0.0))
    if not stores:
        raise RuntimeError("수집된 매장이 없습니다.")
    output_path = Path(args.output)
    write_excel(stores, output_path, collected_at)
    logging.info("완료: %s (%s개 매장)", output_path, len(stores))


if __name__ == "__main__":
    main()
